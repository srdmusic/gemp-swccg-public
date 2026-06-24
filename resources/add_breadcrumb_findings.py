#!/usr/bin/env python3
"""Add a 'Breadcrumb Findings' sheet to the audit Excel + correct the Summary,
based on the backup-doc mining workflow (cross-checked against live code)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PATH = "/Users/steve/gemp-swccg-public/resources/RANDO_BACKUP_AUDIT_2026-06-23.xlsx"
HDR_FILL = PatternFill("solid", fgColor="1F3864"); HDR_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)
ORANGE = PatternFill("solid", fgColor="F8CBAD")  # DOMINATION-RISK
YELLOW = PatternFill("solid", fgColor="FFEB9C")  # UNCODED
BLUE = PatternFill("solid", fgColor="BDD7EE")    # NEEDS-MANUAL-CHECK
GREY = PatternFill("solid", fgColor="E7E6E6")    # SUPERSEDED / DOC-ONLY
GREEN = PatternFill("solid", fgColor="C6EFCE")   # LIVE
VFILL = {"DOMINATION-RISK": ORANGE, "UNCODED": YELLOW, "NEEDS-MANUAL-CHECK": BLUE,
         "SUPERSEDED": GREY, "DOC-ONLY": GREY, "LIVE": GREEN}

# (verdict, finding, in-code, source doc, evidence file:line, recommendation) — sorted by actionability
rows = [
    ("DOMINATION-RISK", "V96 contested-concentrate (+500 flat) vs V67al spread-penalty (power-scaled)", "present", "Audit Findings 1-2", "DeployEvaluator:1832 (+500) vs :3804 (power-scaled); comment :1839-1848 says V96 is meant to beat V67al", "TOP FIX. Make V96 win at ALL power ranges (gate V67al off when contested, or scale V96). Best match for 'spreads out instead of piling on'. VERIFIED."),
    ("UNCODED", "V136 deckShipCount stub (literal 0)", "partial", "V136_DEPLOY_LOG:65", "DeployEvaluator:1766 `0 /* TODO wire */`; CharacterDeploySiteEvaluator:653 shipHeavyDeck=(deckShipCount>=5) never true", "Wire DeckOracle.countShipsInDeck(). §D2 ship-heavy system-cap override is DEAD in both bots. VERIFIED."),
    ("UNCODED", "V136 perSiteEffectActive stub (literal false)", "partial", "V136_DEPLOY_LOG:67", "DeployEvaluator:1767 `false /* TODO wire */`; CharacterDeploySiteEvaluator:468 nbgOverride term dead", "Text-scan active per-site (TDIGWATT) effects, pass real bool. Per-site NBG-penalty override dead. VERIFIED."),
    ("UNCODED", "V136 isAboard stub (always false)", "partial", "V136_HANDOFF:38", "CharacterDeploySiteEvaluator:137 `isAboard=false` never reassigned; used :213 bodyPass", "Implement aboard-ship detection. Pilot-aboard-ship body-count guard wrong. Shared file fixes both bots. VERIFIED."),
    ("DOMINATION-RISK", "V53b Hidden Path +9999 vs V60 landspeed -9999 (implicit precedence)", "present", "Audit Findings 7", "MoveEvaluator:1560 (+9999) vs :1579 (-9999); precedence is execution-order only", "Add explicit precedence so mandatory transit always wins; log a collision warning. High blast (stuck Jedi), lower likelihood."),
    ("DOMINATION-RISK", "V73 Cantina<->Mos Eisley shuttle hardcoded by title", "present", "Audit Findings 27", "MoveEvaluator:2161-2165 .contains('cantina')/('mos eisley'); comment :2127 wants generalization", "Generalize to any two friendly linked drain sites. Overfitting debt; violates search-by-Filter rule. Low gameplay risk."),
    ("DOMINATION-RISK", "Wokling reserve-pull block hardcoded by blueprint ID", "present", "Audit Findings 30", "ActionTextEvaluator:1357-1375 blocks when id=='200_47'; second block :2323+", "Replace ID match with predicate (3-Force Effect pull, turn<=3). Overfitting debt, not a live regression."),
    ("DOMINATION-RISK", "V83/V90 in DeployEvaluator only, no CardSelection mirror", "partial", "K2_HANDOFF_2026-05-22", "DeployEvaluator:1344 V83.1 + :1502 V86.1 guards mitigate; positive logic not mirrored (cf V89-CS at CardSelectionEvaluator:1899)", "LOWER priority now (V83.1/V86.1 removed the false-positive). Add V83-CS/V90-CS only if still observed skipped via CardSelection path."),
    ("UNCODED", "Dojo regression-test framework not built", "absent", "K2_MASTER_HANDOFF §6", "No dojo/, sandbox.sh, live_dojo.py in repo", "SYSTEMIC GAP. No automated guard against a bigger-magnitude rule silently out-scoring an older one. Build before next scoring-tuning wave."),
    ("UNCODED", "Bot-tournament runner scripts not built", "absent", "BOT_TOURNAMENT_HANDOFF:509", "No run/analyze_bot_tournament.py; only build_rando_audit_xlsx.py", "Tooling backlog, not game logic. Build if bot-vs-bot eval needed."),
    ("NEEDS-MANUAL-CHECK", "MLITL weapon-destiny Dark/Light parity unverified", "unknown", "Rule Audit cat 7", "Dark MLITL -6 weapon-destiny modifier; Light mirror not located this pass", "Locate the weapon-destiny tag (CardSelectionEvaluator ~2000-2500 range) and confirm both My Lord Is That Legal sides. Senate deck runs both."),
    ("DOC-ONLY", "V106 status inconsistent (CHANGELOG=removed, code=re-enabled)", "present", "K2_HANDOFF_2026-05-22:376", "ShieldStrategy:575-594 V106 dropped 2026-05-20, RE-ENABLED 2026-06-17 with tighter drain detection", "Code is correct. Update AI_CHANGELOG to record the 2026-06-17 re-enable. Docs-only."),
    ("DOC-ONLY", "V52 SPEND FORCE +300 removed (V_HISTORY stale)", "partial", "Audit Findings 15", "DeployEvaluator:5476 carries 'V52 +300 REMOVED' (V67bk); V52=MOMENTUM only now", "Code correct. Strike the V52 SPEND FORCE entry in AI_VERSION_HISTORY. Docs-only."),
    ("DOC-ONLY", "V29.13 ghost reference to deleted V21", "present", "K2_HANDOFF_2026-05-22:374", "CardSelectionEvaluator:4185 V29.13 correct; V_HISTORY still mentions V21 softening V114 deleted", "Remove the V21 mention from V_HISTORY V29.13 entry. Docs-only."),
    ("SUPERSEDED", "V67y 'duplicate of V29.8' domination (DISPROVED)", "present", "Audit Findings 14", "CardSelectionEvaluator:4688 makes V67y a no-op; V127/V29.8 fully commented, superseded by V153", "DISPROVED by synthesis. No stacking possible. Optional: rename leftover to V67y-SUPERSEDED."),
    ("SUPERSEDED", "V112+V51 double -9999 hard-block (DISPROVED)", "present", "Audit Findings 17", "V51 in DeployEvaluator:2744 is a POSITIVE +500/+600 bonus; V112 is in CardSelectionEvaluator:7755, different path", "DISPROVED. No co-located double block. Disregard."),
    ("SUPERSEDED", "V136 §D disabled in chosenone (DISPROVED)", "present", "changelog agent claim", "Both bots pass live v136ObjRelevant (DeployEvaluator:1762 / chosenone :1726); only deckShipCount/perSiteEffectActive stubbed, identically", "DISPROVED. No rando-vs-chosenone asymmetry. Disregard."),
    ("SUPERSEDED", "BattleEvaluator 'missing ability check' (RESOLVED)", "present", "context.md:411", "BattleEvaluator:45 ABILITY_BATTLE_MAX_POWER_DEFICIT + :185 abilityDiff*2.5 (V164a)", "RESOLVED by V164a. Close the open issue."),
    ("LIVE", "V177 dead-search gate / V178 weapon protection / V179-V182 (Jun 12-14 fixes)", "present", "Rando AI Fixes 4-11", "V177 ActionTextEvaluator:185; V178 CardSelectionEvaluator:4020; V179 DeployPhaseScript:66; V181 CharacterDeploySiteEvaluator:308; V182 DrawEvaluator:228", "All confirmed LIVE. No action."),
    ("LIVE", "V120/V125 weapon-pull block, V82 site-pull +2500, V90/V67aj superseded-by-V136", "present", "K2_HANDOFF / Audit Findings", "V120 ActionTextEvaluator:1832; V82 :4129 (+2500 > V67ai +2000 by design); V90 DeployEvaluator:1779 gated off", "All confirmed LIVE/intended. Add a one-line comment that V82 +2500 > V67ai +2000 is deliberate."),
]

wb = openpyxl.load_workbook(PATH)
if "Breadcrumb Findings" in wb.sheetnames:
    del wb["Breadcrumb Findings"]
w = wb.create_sheet("Breadcrumb Findings", 1)  # second tab
w["A1"] = "Breadcrumb Findings — backup docs cross-checked against LIVE code (2026-06-23)"; w["A1"].font = TITLE_FONT
w["A2"] = ("NOT a merge list. The current code is the most complete version; these are issues the backup docs/audits documented that my V-tag-presence pass could not see. "
           "Verdicts: DOMINATION-RISK = present but possibly out-scored by a later rule (the real shape of 'dumb moves I thought we fixed'); UNCODED = documented, dead/stubbed; "
           "SUPERSEDED = already fixed or a disproved agent claim; DOC-ONLY = changelog text reconciliation; LIVE = confirmed working.")
w["A2"].alignment = WRAP; w.merge_cells("A2:G2"); w.row_dimensions[2].height = 54
cols = ["Verdict", "Finding", "In live code", "Source doc", "Evidence (file:line)", "Recommendation"]
for j, c in enumerate(cols):
    cell = w.cell(row=4, column=j + 1, value=c); cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = WRAP; cell.border = THIN
for i, r in enumerate(rows):
    for j, v in enumerate(r):
        cell = w.cell(row=5 + i, column=j + 1, value=v); cell.alignment = WRAP; cell.border = THIN
    w.cell(row=5 + i, column=1).fill = VFILL.get(r[0], GREY)
    w.row_dimensions[5 + i].height = 50
for col, wd in zip("ABCDEF", [20, 44, 12, 24, 50, 56]):
    w.column_dimensions[col].width = wd
w.freeze_panes = "A5"

# Correct the Summary bottom line
s = wb["Summary"]
s.cell(row=10, column=1, value="REVISION (breadcrumb audit)").font = Font(bold=True, color="C00000")
s.cell(row=10, column=2, value=("Nothing to MERGE from backups still holds (current code is the most complete version). BUT the backup docs surfaced issues a V-tag-presence grep cannot see: "
    "1 confirmed silent-domination (V96 +500 flat loses to V67al power-scaled spread penalty at high power -> Rando declines to pile on a contested site, matching 'spreads instead of piling on'); "
    "3 dead V136 stubs (deckShipCount/perSiteEffectActive/isAboard fed literal 0/false, so 3 deploy overrides never fire); a V53b/V60 precedence collision; and the unbuilt dojo regression harness (the systemic guard against this whole class). "
    "5 items warrant code work; see the Breadcrumb Findings tab. My earlier 'nothing missing' was V-tag-presence only."))
s.cell(row=10, column=2).alignment = WRAP; s.row_dimensions[10].height = 96

wb.save(PATH)
print("UPDATED", PATH); print("sheets:", wb.sheetnames)
